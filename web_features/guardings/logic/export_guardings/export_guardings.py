from APIs.TalpiotAPIs.Tasks.guarding.guarding_week import GuardingWeek
from APIs.TalpiotAPIs.Tasks.guarding.guarding_week import GuardingDay
from APIs.TalpiotAPIs.Tasks.task import Task
from APIs.TalpiotAPIs.Tasks.task_type import TaskType
from APIs.TalpiotAPIs.User.user import User
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Font
# import excel2img
from PIL import Image
import datetime
import os
from typing import List
import random

from APIs.TalpiotAPIs.mahzors_utils import get_mahzor_color


class ExportGuardingWeek:
    DAYS = ["יום ראשון", "יום שני", "יום שלישי", "יום רביעי", "יום חמישי"]
    TITLE_COLOR = '00AAFF'
    BLACK = '000000'
    TOP_BORDER = Border(top=Side(border_style='thin', color=BLACK),
                        left=Side(border_style='thin', color=BLACK),
                        right=Side(border_style='thin', color=BLACK))
    SIDE_BORDER = Border(left=Side(border_style='thin', color=BLACK),
                         right=Side(border_style='thin', color=BLACK))
    BOTTOM_BORDER = Border(bottom=Side(border_style='thin', color=BLACK),
                           left=Side(border_style='thin', color=BLACK),
                           right=Side(border_style='thin', color=BLACK))
    SEPARATOR = Border(top=Side(border_style='thin', color=BLACK))

    def __init__(self, guarding_week: GuardingWeek, image_name: str) -> None:
        """
        Constructor of ExportGuardingWeek
        :param guarding_week: a GuardingWeek object
        :param image_name: the file name for saving the image
        """
        self.guarding_week = guarding_week
        self.is_updated = False
        self.guarding_photo = None
        self.image_name = image_name
        self.OFFSET = 2
        # List of the first columns
        # Reverse order because that the columns are Left to Right
        self.COLUMNS = 'ABCDEFGIJKLMNOPQRTSUVWXYZ'[len(self.guarding_week) - 1 + self.OFFSET::-1]
        self.FONT = "Gisha"

    def update_guarding_week(self, guarding_week: GuardingWeek) -> None:
        """
        Update the saved guarding week and mark that the current photo isn't updated anymore
        :param guarding_week: the new guarding week object
        :return: None
        """
        self.is_updated = False
        self.guarding_week = guarding_week

    def get_guarding_photo(self) -> str:
        """
        Generate a relevant guarding photo if needed, and return the relevant photo
        :return: the up-to-date guarding photo file name
        """
        if not self.is_updated:
            file_name = str(random.randint(0, 2**20))
            self.__generate_excel(file_name+ '.xlsx')
            self.is_updated = True
            self.__convert_to_image(file_name + '.xlsx', file_name + '.png')
            os.remove(file_name + '.xlsx')
            # Upload to drive
            self.guarding_photo = file_name + '.png'
            return file_name + '.png'

    def __generate_excel(self, file_name) -> None:
        """
        Private method for generating an excel file from the guarding week
        ;:param: file_name - the name of the excel file
        :return: None
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Guarding Week'
        for col in self.COLUMNS:
            sheet.column_dimensions[col].bestFit = True

        # Calculate the maximum number of people in each task
        people_for_tasks = []
        max_tasks = max([len(day.guardings) for day in self.guarding_week.days])
        for task_num in range(max_tasks):
            people_for_tasks.append(max(day.guardings[task_num].task_type.required_people
                                        for day in self.guarding_week.days if len(day.guardings) > task_num))

        self.__add_titles(sheet)
        self.__add_tasks_details(sheet, people_for_tasks)
        for i, day in enumerate(self.guarding_week.days):
            column = self.COLUMNS[i + self.OFFSET]
            self.__add_day_tasks(sheet, day, column, 3, people_for_tasks)

        for col in self.COLUMNS:
            for row in range(1, sum(people_for_tasks) + self.OFFSET + 1):
                sheet[f"{col}{row}"].font = Font(name=self.FONT)

        workbook.save(filename=file_name)
        workbook.close()

    def __add_titles(self, sheet: Workbook.worksheets) -> None:
        """
        Private method for adding titles to the guardings excel
        :param sheet: the sheet where the guarding list is saved
        :return: None
        """
        day_cells = [f"{col}1" for col in self.COLUMNS[self.OFFSET:len(self.guarding_week.days) + self.OFFSET]]
        date_cells = [f"{col}2" for col in self.COLUMNS[self.OFFSET:len(self.guarding_week.days) + self.OFFSET]]
        title_fill = PatternFill(fgColor=Color(rgb=self.TITLE_COLOR), fill_type='solid')
        for i, cell in enumerate(day_cells):
            sheet[cell] = self.DAYS[i]
            sheet[cell].fill = title_fill
            sheet[cell].alignment = Alignment(horizontal='center')
            sheet[cell].border = self.TOP_BORDER
        for i, cell in enumerate(date_cells):
            sheet[cell] = self.guarding_week.days[i].date.strftime("%d/%m/%Y")
            sheet[cell].fill = title_fill
            sheet[cell].alignment = Alignment(horizontal='center')
            sheet[cell].border = self.BOTTOM_BORDER

    def __add_day_tasks(self, sheet: Workbook.worksheets, task_day: GuardingDay, column: str,
                        first_row: int, people_for_tasks: List[int]) -> None:
        """
        Private method for adding the guarding list of a specific day to the excel
        :param sheet: the sheet where the guarding list will bw saved
        :param task_day: the GuardingDay object which holds the data to save
        :param column: the column of the day in the excel file
        :param first_row: the first row to start writing from
        :param: people_for_tasks: a list of the maximum number of people in each task
        :return: None
        """
        for task_index, task in enumerate(task_day.guardings):
            current_row = first_row + sum(people_for_tasks[:task_index])
            for user_index, task_user in enumerate(task.assignment):
                cell = f"{column}{current_row}"
                sheet[cell] = task_user.name
                sheet[cell].border = self.SIDE_BORDER if user_index != 0 else self.TOP_BORDER
                sheet[cell].alignment = Alignment(horizontal='center')
                mahzor_color = Color(rgb=get_mahzor_color(task_user.mahzor).replace('#', ''))
                cell_fill = PatternFill(fgColor=mahzor_color, fill_type='solid')
                sheet[cell].fill = cell_fill
                current_row += 1
            sheet[f"{column}{current_row - 1}"].border = self.BOTTOM_BORDER

    def __add_tasks_details(self, sheet: Workbook.worksheets, people_for_tasks: List[int]):
        """
        Add the time of each task and its type
        :param sheet: the sheet of the guarding list
        :param people_for_tasks: list of the maximum number of people in each task
        :return: None
        """
        type_column = self.COLUMNS[0]
        hours_column = self.COLUMNS[1]
        day_max_tasks = max([i for i in range(len(self.guarding_week.days))],
                            key=lambda i: len(self.guarding_week.days[i].guardings))
        title_fill = PatternFill(fgColor=Color(rgb=self.TITLE_COLOR), fill_type='solid')
        current_row = 3
        for index_task, task in enumerate(self.guarding_week.days[day_max_tasks].guardings):
            sheet[f"{type_column}{current_row}"] = task.task_type.description
            sheet[f"{type_column}{current_row}"].border = self.SEPARATOR
            sheet[f"{type_column}{current_row}"].fill = title_fill
            duration_str = f"{task.start_time.strftime('%H:%M')} - {task.end_time.strftime('%H:%M')}"
            sheet[f"{hours_column}{current_row}"] = duration_str
            sheet[f"{hours_column}{current_row}"].border = self.SEPARATOR
            sheet[f"{hours_column}{current_row}"].fill = title_fill
            for row in range(current_row, current_row + people_for_tasks[index_task]):
                sheet[f"{type_column}{row}"].fill = title_fill
                sheet[f"{hours_column}{row}"].fill = title_fill

            current_row += people_for_tasks[index_task]

    def __convert_to_image(self, excel_file, image_file):
        """
        Convert the excel file to an image file
        Supports PNG and BMP file formats
        :param: excel_file - the name of the temp excel file
        :param: image_file - the name of the temp image file
        :return: None
        """
        # excel2img.export_img(excel_file, image_file, 'Guarding Week', None)
        pass


if __name__ == '__main__':
    user1 = User(name="חניך אקראי", mahzor=43)
    user2 = User(name="פלוני אלמוני", mahzor=42)
    user3 = User(name="ישראל ישראלי", mahzor=41)

    now = datetime.datetime.now()
    day = GuardingDay(date=now, guardings=[Task(start_time=now, end_time=now, assignment=[user1, user2],
                                                task_type=TaskType(description="שמירת יום",
                                                                   required_people=2)),
                                           Task(start_time=now, end_time=now, assignment=[user1],
                                                task_type=TaskType(description="שמירת לילה",
                                                                   required_people=1)),
                                           Task(start_time=now, end_time=now, assignment=[user2, user1, user3],
                                                task_type=TaskType(description="שמירת יום",
                                                                   required_people=3))])
    day2 = GuardingDay(date=now, guardings=[Task(start_time=now, end_time=now, assignment=[user1, user2],
                                                task_type=TaskType(description="שמירת יום",
                                                                   required_people=2)),
                                           Task(start_time=now, end_time=now, assignment=[user1, user2],
                                                task_type=TaskType(description="שמירת לילה",
                                                                   required_people=2))])
    week = GuardingWeek(days=[day, day2, day, day, day])
    export = ExportGuardingWeek(week, 'guarding.png')
    export.get_guarding_photo().show()

