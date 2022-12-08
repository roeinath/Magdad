# !!!!!!!!!!!! library 'openpyxl' must be installed !!!!!!!
import openpyxl


class ExcelParser:
    @staticmethod
    def load_workbook(path: str) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(path)
        return wb

    def __init__(self, path_for_file: str):
        self.path = path_for_file
        self.wb = ExcelParser.load_workbook(self.path)

    @property
    def sheets(self):
        return self.wb.worksheets

    def get_data_from_sheet(self, sheet_index):
        """Get data from a specific sheet in the file."""
        sheet = self.sheets[sheet_index]
        data = []

        # Iterating over all 'interesting' rows
        for row in range(2, sheet.max_row + 1):
            row_data = {}
            # Iterating over all 'interesting' columns
            for col in range(2, sheet.max_column + 1):
                # Retrieving header and value
                category = sheet.cell(row=1, column=col).value
                value = sheet.cell(row=row, column=col).value

                # Adding data to the specific row data
                row_data[category] = value

            data.append(row_data)
        return data

    def get_data(self):
        data = {}
        for sheet_number in range(len(self.sheets)):
            data[self.sheets[sheet_number].title] = self.get_data_from_sheet(sheet_number)
        return data

    def get_first_col(self, sheet_index):
        sheet = self.sheets[sheet_index]
        return [sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)]

# EXAMPLE OF USAGE
if __name__ == '__main__':
    my_excel_parser = ExcelParser('filename.xlsx')

    # Data from the 2nd sheet
    print(my_excel_parser.get_data_from_sheet(2))

    # Data from all sheets
    print(my_excel_parser.get_data())
